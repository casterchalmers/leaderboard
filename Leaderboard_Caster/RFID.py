#CASTER LEADERBOARD
#VERSION 0.0.1
#Axel Lindmark


from tkinter import *
from screeninfo import get_monitors
from openpyxl import load_workbook
import os
import sys
import subprocess
import time


#Gets scrren information eg screen seize
screen_width = get_monitors()[0].width
screen_height = get_monitors()[0].height

#Starts a new window
window = Tk()

#Sets the seize of window
window.geometry(f'{screen_width}x{screen_height}')
window.attributes("-fullscreen", True)
window.configure(bg="Black")

#Skapar anv. inputs för rfid och anv. namn.
user_input = Entry()
user_name_input = Entry()
#Skapar text output för anv. t.ex "Ditt RFID finns redan!"
answer = Label()

while True:
    try:
        database = load_workbook(filename = r"readme2.xlsx")
    except ValueError:
        time.sleep(1)
    else:
        database = load_workbook(filename=r"readme2.xlsx")
        break

#Funktionen som  tar emot RFID info, körs då anv. trycker på "Input RFID". Funktionen kollar om RFID infon är giltig, eg om den är längre än 4 char. Om anv. har ett
#tillräckligt långt RFID tag så kollar funktionen om det redan finns en xlsx fil att lägga till RFID/anv. namn i. Om det ej finns en sådan fil skapas en sådan fil, om
#det redan finns en fil så öpnnas denna. Oavsett fall så körs funktionen check_if_exist. Beroende på svaret från check_if_exist (1 = anv. finns redan i xlsx filen, 2 =
#anv. finns ej) så får anv. antigen tillgång till en knapp för att lägga till sitt namn, med new_name, eller, om anv. redan finns, så görs något som jag ej definierat än
def get_RFID():

    rfid = user_input.get()

    file_rfid = open("rfid_bin.txt", "w")
    file_rfid.write(rfid)
    file_rfid.close()

    if len(rfid) < 4:

        answer = Label(text="Thats probably not a RFID tag!")
        answer.config(bg="black", fg="white", font=("Arial", 30), relief = "flat")
        answer.place(x=0, y=110)
        user_input.delete(0, END)

        answer.after(5000, answer.destroy)

    else:

        user_input.delete(0, END)
        user_input.insert(0, "")

        is_exist = os.path.exists(r"readme2.xlsx")

        if is_exist == True:
            global database
            user_status = check_if_exist(database, rfid)

            if user_status[0] == 2:
                #print("User status is 2 in get_rfid [get_rfid]")

                excel_func(database, rfid)

                button_name.place(x=900, y=805)

            elif user_status[0] == 1:
                #print("User status is 1 in get_rfid [get_rfid]")
                find_name(database,user_status)

#funktionen excel_func loopar igenom "hela" excel arket och hittar vart den första tomma platsen är. Detta görs genom att loopa genom varje rad och
#kolla vart "none" fås. Detta läget sparas som first_empty_cell. På den platsen läggs sedan den inmatede RFID taggen.
def excel_func(database, rfid):

    sheet = database.active
    i = 0

    for i in range(1, 5000):

        value_input = sheet.cell(row=i, column=1)

        if value_input.value == None:
            #print("broke loop in excel_func, first empty position found! [excel_func]")
            break

        else:
            #har inget egentligt syfte
            cell_val = sheet.cell(row=i, column=1)
            #print(cell_val.value)

    actual_cell = i -1
    first_empty_cell = i


    time = sheet.cell(row = first_empty_cell, column = 3)
    time.value = "99:99:999"
    #time.value = "10000000000000000000"

    input_rfid = sheet.cell(row = first_empty_cell, column = 1)
    input_rfid.value = rfid

    while True:
        try:
            database.save(filename=r"readme2.xlsx")
        except ValueError:
            time.sleep(1)
        else:
            database.save(filename=r"readme2.xlsx")

            break

    return first_empty_cell

#check_if_exist loopar igenom "alla" rader i xlsx arket och kollar om den kan hitta det inmatade rfid taggen. Om namnet hittas sätts user_status som 1 och om namnet inte hittas
#sättas den som 2.
def check_if_exist(database, rfid):

    sheet = database.active

    where_found_rfid = 0

    user_status = 0

    for i in range(1, 5000):

        value_input = sheet.cell(row=i, column=1)

        if value_input.value == None:
            #print("loop found empty cell in check_if_exist, this means nothing")
            break

        elif value_input.value == rfid:

            user_status = 1
            where_found_rfid = i

            break


        elif value_input.value != rfid:
            user_status = 2


    #print(user_status)

    if user_status == 1:

        answer2 = Label(text="RFID is registred!")
        answer2.config(bg="black", fg="white", font=("Arial", 30), relief="flat",justify=CENTER, width=50)
        answer2.pack()
        user_input.delete(0, END)
        answer2.after(3000, answer2.destroy)

    elif user_status == 2:
        #print("has run the else")
        answer3 = Label(text="RFID not registred, input name below!")
        answer3.config(bg="black", fg="white", font=("Arial", 30), relief="flat")
        answer3.place(x=0, y=110)
        user_input.delete(0, END)
        button_RFID.after(10, button_RFID.destroy)
        answer3.after(3000, answer3.destroy)

    return user_status, where_found_rfid


#new_name visar endast knappen för att lägga till namn
def new_name():

    sheet = database.active

    new_name = user_name_input.get()

    user_name_input.delete(0, END)
    user_name_input.insert(0, "")

    i = 0

    for i in range(1, 5000):

        value_input = sheet.cell(row=i, column=1)

        if value_input.value == None:
            #print("broke loop in excel_func, first empty position found! [excel_func]")
            break

        else:
            # har inget egentligt syfte
            cell_val = sheet.cell(row=i, column=1)

    first_empty_cell = i -1

    if new_name == "":
        answer4 = Label(text="Enter a name!")
        answer4.config(bg="black", fg="white", font=("Arial", 30), relief="flat")
        answer4.place(x=0, y=910)
        answer4.after(3000, answer4.destroy)

    else:

        input_name = sheet.cell(row=first_empty_cell, column=2)
        input_name.value = new_name

        while True:
            try:
                database.save(filename=r"readme2.xlsx")
            except ValueError:
                time.sleep(1)
            else:
                database.save(filename=r"readme2.xlsx")
                break

        file = open("name_bin.txt", "w")
        file.write(input_name.value)
        file.close()


        #program-ending---------------------------------------------------------
        user_input.destroy()
        button_RFID.after(10, button_RFID.destroy)

        user_name_input.destroy()
        button_name.after(10, button_name.destroy)

        answer5 = Label(text="Logging in...")
        answer5.config(bg="black", fg="white", font=("Arial", 30), relief="flat")
        answer5.pack()


        window.quit()
        window.destroy()

        subprocess.call([r'fetch_race_handler.bat'])

        sys.exit()



def find_name(database, user_status):

    sheet = database.active

    value_name = sheet.cell(row=user_status[1], column=2)

    words = "Welcome, " + value_name.value

    answer3 = Label(text=words)
    answer3.config(bg="black", fg="white", font=("Arial", 30), relief="flat", justify=CENTER)
    answer3.pack()
    user_input.delete(0, END)

    answer3.after(3000, answer3.destroy)

    file = open("name_bin.txt", "w")
    file.write(value_name.value)
    file.close()

    #program ending-------------------------------------------------------
    user_input.destroy()
    button_RFID.after(10, button_RFID.destroy)

    user_name_input.destroy()
    button_name.after(10, button_name.destroy)

    answer5 = Label(text="Logging in...")
    answer5.config(bg="black", fg="white", font=("Arial", 30), relief="flat")
    answer5.pack()

    print("running prog")
    window.quit()
    window.destroy()

    subprocess.call([r'fetch_race_handler.bat'])

    sys.exit()



#config för input fönster

user_input.config(bg='Grey', fg='White', font=('Arial', 60), show='*', relief = "flat")
user_input.config(state=NORMAL)

user_name_input.config(bg='Grey', fg='White', font=('Arial', 60), relief = "flat")
user_name_input.config(state=NORMAL)

#definierar knapp
button_RFID = Button(text="Input RFID", command=get_RFID, height=1, activebackground='Green', bg='Grey', fg='white', font=('Arial', 30))
button_name = Button(text="Input Name", command=new_name, height=1, activebackground='Green', bg='Grey',fg='white', font=('Arial', 30))

#lägger in input delen i fönstret
user_input.place(x=0, y=0)
user_name_input.place(x=0, y=800)

button_RFID.place(x=900, y=6)


window.mainloop()
