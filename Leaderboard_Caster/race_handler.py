#CASTER LEADERBOARD
#VERSION 0.0.1
#Axel Lindmark

from tkinter import *

from screeninfo import get_monitors
from openpyxl import load_workbook
import sys
import subprocess
import linecache
import re
import time
from datetime import *

laplog_path_TA = "D:\Studier\Chalmers\Caster\Leaderboard_Caster\laplog.txt"
#laplog_path_UT = "L:\dump\laplog.txt" #Vad gör denna?
laplog_path = laplog_path_TA

#Gets scrren information eg screen seize
screen_width = get_monitors()[0].width
screen_height = get_monitors()[0].height

#Starts a new window
window = Tk()

#Sets the seize of window
window.geometry(f'{screen_width}x{screen_height}')
window.attributes("-fullscreen", True)
window.configure(bg="Black")

#------------------------------------------------------------

#cursor_path = "@busy.cur"
#window["cursor"] = cursor_path

#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
def start_race():

    time_today = str(datetime.today())

    year = time_today[0:4]
    month = time_today[5:7]
    day = time_today[8:10]
    time = time_today[11:13] + time_today[14:16] + time_today[17:19]

    time_ref = int(f"{year}{month}{day}{time}")

    return time_ref

def kill_ui():
    welcome.destroy()
    button_race.destroy()
    button_finsih_race.pack()


time_ref = start_race()


def finish_race():
#------------------------------------------------------------
    #cursor_path = "@busy.cur"
    #window["cursor"] = cursor_path
    wait = Label(window,text="Please wait(30sec)....",bg="black", fg="white", font=("Arial", 30), relief="flat", justify=CENTER)
    wait.pack()

    window.update()
    window.after(3000) #30000

    file = open(laplog_path, "r")
    #print(time_ref)

    laplog_text = file.read()
    nlines = laplog_text.count('\n')

    for x in reversed(range(nlines)):
        current_line = linecache.getline(laplog_path, x)
        print(current_line)
        if "[RACE_END]" in current_line and x>36300:

            last_date_finsish = current_line[0:21]
            last_date_line_finish = x
            #print(current_line)

            year = last_date_finsish[6:10]
            month = last_date_finsish[3:5]
            day = last_date_finsish[0:2]
            time = last_date_finsish[12:14] + last_date_finsish[15:17] + last_date_finsish[18:20]

            time_ref_race = int(f"{year}{month}{day}{time}")
            #print(f"{time_ref_race} time_ref_race")
            #print(f"{time_ref} ref_time")

            if "porsche_911_gt3" in current_line and "autodromo_de_lunyachamps" in current_line and time_ref_race>time_ref:
            #"last_date_finish" != last_date and "laste_date_line_finish" != last_date_line and

                new_time_list = re.findall(r'(\S+):(\S+):(\S+)', current_line)
                new_time_tup = new_time_list[1]
                new_time = new_time_tup[0] + ":" + new_time_tup[1] + ":" + new_time_tup[2]

                #print(new_time)
                #input_excel(new_time)

                file.close()

                break
        else:
            window.quit()
            window.destroy()
            subprocess.call([r'fetch_RFID.bat'])
            sys.exit()



#---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------



def input_excel(new_time):
    file_rfid = open("rfid_bin.txt", "r")
    rfid_user = file_rfid.read()
    file_rfid.close()

    while True:
        try:
            database = load_workbook(filename=r"readme2.xlsx")
        except ValueError:
            time.sleep(1)
        else:
            database = load_workbook(filename=r"readme2.xlsx")
            break


    sheet = database.active
    global user_name


    for i in range(1,5000):

        cell_rfid = sheet.cell(row = i+1, column=1)

        if cell_rfid.value != None and rfid_user in cell_rfid.value:

            new_time_edit = new_time.replace(":", "")
            new_time_float = int(new_time_edit)

            old_time = sheet.cell(row = i+1, column=3)
            old_time = old_time.value

            old_time_edit = old_time.replace(":", "")
            old_time_float = int(old_time_edit)

            if new_time_float < old_time_float:

                cell_time = sheet.cell(row = i+1, column=3)
                cell_time.value = new_time

                while True:
                    try:
                        database.save(filename=r"readme2.xlsx")
                    except ValueError:
                        time.sleep(1)
                    else:
                        database.save(filename=r"readme2.xlsx")
                        break

                window.quit()
                window.destroy()
                subprocess.call([r'fetch_RFID.bat'])
                sys.exit()

            else:
                window.quit()
                window.destroy()
                subprocess.call([r'fetch_RFID.bat'])
                sys.exit()


file = open("name_bin.txt", "r")
user_name = file.read()
file.close()



words = "Welcome, " + user_name

distance = Label(text = "A")
distance.config(bg="black", fg="black", font=("Arial", 200), relief="flat", justify=CENTER)
distance.pack()


welcome = Label(text=words)
welcome.config(bg="black", fg="white", font=("Arial", 30), relief="flat", justify=CENTER)
welcome.pack()


button_race = Button(text="Start Race", command=kill_ui, height=1, bg='Green', fg='white', font=('Arial', 30))
button_race.pack()


button_finsih_race = Button(text="Finish Race", command=finish_race, height=1, bg='Red', fg='white', font=('Arial', 30), activebackground="white")

window.mainloop()



subprocess.call([r'fetch_RFID.bat'])








